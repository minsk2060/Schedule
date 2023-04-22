import webbrowser, schedule, time
from pyautogui import hotkey
from openpyxl import load_workbook
from pywinauto.application import Application


print("РАБОТАЕТ УПРАВЛЕНИЕ ОБОРУДОВАНИЕМ ПО РАСПИСАНИЮ, НЕ ЗАКРЫВАЙТЕ ЭТО ОКНО")

schedule_book="C://Users/BMS/projects/schedules/Расписание.xlsm"
tasks=[]  # Список tasks сотоит из списков single и содержит все прочитанные расписания, даже пустые
single=[] # Список single содержит одну строку расписания: установка такая-то сделать то-то в такое-то время


# Очистка списка single и пополнение списка tasks
def refresh():
    copysingle = single.copy()   # Создать копию списка single
    tasks.append(copysingle)     # Добавить в конец списка tasks эту копию
    single.clear()               # Очистить список single

# Отправка запроса в браузер
def turn(get_plant, par):
    PREFX = "http://192.168.250.50/ajaxjson/bac/setValue?pid=85&oid="
    app = Application(backend="uia").connect(title_re=".*Microsoft\u200b Edge", timeout=10)
    app.window().set_focus()
    webbrowser.open_new_tab(PREFX + get_plant + par)
    time.sleep(3)
    hotkey("ctrl", "w")


# Reading the schedules and matching the list of them
def runschedule():
    workbook = load_workbook(schedule_book)
    for j in range(53, 383, 10):                                                 # Цикл чтения всех значений с 53 до 383 строки с шагом 10
        for k in range(7):                                                       # Цикл чтения по строкам
            single.append(str(workbook.active.cell(row=j,     column=4).value))  # Читать код установки для запроса


            # Читать значения для "старт" из первого массива расписаний
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))  # Читать день недели
            single.append(str(workbook.active.cell(row=j+1+k, column=5).value))  # Читать время действия старт
            single.append(str(workbook.active.cell(row=j+1+k, column=3).value))  # Читать само действие
            refresh()                                                            # Очистить список single и пополнить список списков tasks


            # Читать значения для "стоп" из первого массива расписаний
            single.append(str(workbook.active.cell(row=j,     column=4).value))  # Читать код устнановки для запроса
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))  # Читать действие
            single.append(str(workbook.active.cell(row=j+1+k, column=6).value))  # Читать время действия стоп


            # Отдельная проверка для расписаний больших вентустановок ПВ-1.1, ПВ-2.1
            driers = str(workbook.active.cell(row=j, column=4).value)            # Читать код установки
            if driers == "79691782&did=33556432" or driers == "79691777&did=33555432": # Провериьт на совпадение с кодом ПВ-1.1 или ПВ-2.1
                single.append("5") # Присвоить 5 (параметр останова в запросе 5 для ПВ-1.1, ПВ-2.1)
            else:
                single.append("0") # Присвоить 0  (параметр останова в запросе 0 для всех остальных установок
            driers = ""            # Очистить переменную driers
            refresh()              # Очистить список single и пополнить список списков tasks


            # Читать значения для "старт" из второго массива расписаний
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=8).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=10).value))
            refresh() # Очистить список single и пополнить список списков tasks


            # Читать значения для "старт" из второго массива расписаний
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=9).value))
            driers = str(workbook.active.cell(row=j, column=4).value)


            # Отдельная проверка для расписаний больших вентустановок ПВ-1.1, ПВ-2.1
            if driers == "79691782&did=33556432" or driers == "79691777&did=33555432":
                single.append("5")
            else:
                single.append("0")
            driers = ""
            refresh() # Очистить список single и пополнить список списков tasks


    # Удаление пустых строк из списка расписаний tasks
    cleartasks = tasks.copy()          # Сделать копию списка tasks, который сейчас содержит все строки расписаний
    for t in range(len(tasks)):        # Пройтись в цикле по всем компонентам списка списков
        if tasks[t][2] == "None":      # Если время не задано
            cleartasks.remove(tasks[t])# Удалить воженный пустой список
    schedule.clear()                   # Очистить предыдущее расписание

    # keyboard()
    # if keyboard() == "ru":
    #     hotkey("alt", "shift")

    for i in range(len(cleartasks)):
        # выполнить распсиание из всего списка , где:
        # cleartasks[i][1]   - день недели
        # cleartasks[i][2]   - время
        # cleartasks[i][0]   - код установки-
        # cleartasks[i][3]   - действие
        exec(f"""schedule.every().{cleartasks[i][1]}.at('{cleartasks[i][2]}').do(turn, '{(cleartasks[i][0])}','&vid=17&value={cleartasks[i][3]}')""")

    schedule.every(10).minutes.do(runschedule) # Каждые 10 минут читать расписание
    cleartasks.clear()                         # Очистить список списков с расписаниями, где нет пустых строк
    tasks.clear()                              # Очистить список списков с расписаниями, где есть пустые строки

runschedule() # Запуск функции для старта программы


while True:
    schedule.run_pending()
    time.sleep(1)






