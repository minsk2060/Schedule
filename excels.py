from openpyxl import load_workbook
from datetime import datetime
from logs import sort, Textjob
from plants import driers
from TelegramBot import to_telegram
from ViberSet import to_viber

schedule_book = "./excel/Расписание.xlsm"
status_book   = "./excel/Cостояние.xlsx"
single = []


def refresh(tasks):
    """
    refresh()        - очистка списка single и пополнение списка tasks
    tasks            - полный список расписаний, в т.ч. пустых
    single           - по итогу список с одной задачей
                       (установка, день недели, что сделать, во сколько сделать)
    replace()        - исправление не корректного чтения
    """
    single[0].replace(" ", "")
    tasks.append(single.copy())
    single.clear()

def refreeze(row, lis, book):
    """
    refreeze()       - чтение кода установки из файла excel
    row              - номер ряда в файле excel
    lis              - список single
    book             - рабочая книга Расписание.xlsm
    plancod          - прочтенное из файла excel значение кода установки
    """
    plancod = (str(book.active.cell(row=row, column=4).value))
    lis.append(plancod)
    return plancod

def drier(row, lis, book):
    """
    drier()          - проверка не является ли установка осушителем
                       (имеет нетиповые параметры управления)
    driers           - коды установок, где стопу соответствует параметр "5" , а не "0"
    row              - номер ряда в файле excel
    lis              - список single
    book             - рабочая книга Расписание.xlsm
    """
    dry = str(book.active.cell(row=row, column=4).value)
    if dry in driers:
        lis.append("5")
    else:
        lis.append("0")

def readschedule(tasks):
    """
    readschedule()   - чтение расписания из excel файла принимает пустой список tasks, возвращает заполненный список
    tasks            - полный список расписаний, в т.ч. пустых
    workbook         - работа с рабочей книгой excel
    refresh()        - предобработка заданий с расписанием
    """
    workbook = load_workbook(schedule_book)
    for j in range(53, 383, 10):
        for k in range(1, 8):
            refreeze(j, single, workbook)
            for s in [2, 5, 3]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 6]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            drier(j, single, workbook)
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 8, 10]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 9]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            drier(j, single, workbook)
            refresh(tasks)
    return tasks

def writestatus(i, plant, alarm, column):
    """
    writestatus()  - запись сведений об аварии в файл Состояние.xlsx
    i              - сдвиг номера строки для внесения записей в файл Состояние.xlsx
    plant          - имя установки, заносимое в файл Состояние.xlsx (например "ПВ-1.6")
    alarm          - запись об аварии, заносимая в файл Состояние.xlsx (например "Авария класса А")
    column         - номер колонки для записи в файл Состояние.xlsx
    statusbook     - объект для работы с файлом Состояние.xlsx
    date_now       - текущая дата, заносимая в файл Состояние.xlsx
    time_now       - текущее время, заносимая в файл Состояние.xlsx
    alarm_happen   - сторка типа "04-06-2023  11:00  ПВ-2.8   Авария класса А"
    makelog()      - занесение события об аварии в лог файл alllogs.txt
    sort()         - занесение события об аварии в лог файл log_scheduling.txt
    to_telegram()  - отправка сообщения об аварии в чат бот в Телеграм
    to_viber()     - отправка сообщения об аварии в чат бот в Viber
    """
    statusbook = load_workbook(status_book)
    date_now = datetime.now().strftime("%d-%m-%Y")
    time_now = datetime.now().strftime("%H:%M")
    statusbook.active.cell(row=3+i, column=1).value = date_now
    statusbook.active.cell(row=3+i, column=2).value = time_now
    statusbook.active.cell(row=3+i, column=3).value = plant
    if statusbook.active.cell(row=3+i, column=column).value != alarm:
        alarm_happen = [f"{date_now}  ", f"{time_now}  ", f"{plant}  ",  alarm]
        m = Textjob(Textjob.pathall, "a")        #new
        m.makelog(alarm_happen, "\n")            #new
       # logall(alarm_happen)                    #old
        sort()
        logmsg = "\n".join(alarm_happen[2:])
        to_telegram(logmsg)
        to_viber(logmsg)
    statusbook.active.cell(row=3 + i, column=column).value = alarm
    statusbook.save(status_book)



