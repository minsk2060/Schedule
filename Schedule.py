import webbrowser, schedule, time
from pyautogui import hotkey
from openpyxl import load_workbook
from pywinauto.application import Application
import ctypes

print("РАБОТАЕТ УПРАВЛЕНИЕ ОБОРУДОВАНИЕМ ПО РАСПИСАНИЮ, НЕ ЗАКРЫВАЙТЕ ЭТО ОКНО")

schedule_book="C://Users/BMS/projects/schedules/Расписание.xlsm"
tasks=[]
single=[]


# Refreshing the list
def refresh():
    copysingle = single.copy()
    tasks.append(copysingle)
    single.clear()
def turn(get_plant, par):
    # u = ctypes.windll.LoadLibrary("user32.dll")
    # pf = getattr(u, "GetKeyboardLayout")
    # if hex(pf(0)) == '0x4190419':
    #     hotkey('alt','shift')
    # # if hex(pf(0)) == '0x4090409':
    #     return 'en'

    PREFX = "http://192.168.250.50/ajaxjson/bac/setValue?pid=85&oid="
    app = Application(backend="uia").connect(title_re=".*Microsoft\u200b Edge", timeout=10)
    app.window().set_focus()
    webbrowser.open_new_tab(PREFX + get_plant + par)
    time.sleep(3)
    hotkey("ctrl", "w")


# Reading the schedule and matching the list of them
def runschedule():
    workbook = load_workbook(schedule_book)
    for j in range(53, 383, 10):
        for k in range(7):
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=5).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=3).value))
            refresh()
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=6).value))
            driers = str(workbook.active.cell(row=j, column=4).value)
            if driers == "79691782&did=33556432" or driers == "79691777&did=33555432":
                single.append("5")
            else:
                single.append("0")
            driers = ""
            refresh()
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=8).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=10).value))
            refresh()
            single.append(str(workbook.active.cell(row=j,     column=4).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=2).value))
            single.append(str(workbook.active.cell(row=j+1+k, column=9).value))
            driers = str(workbook.active.cell(row=j, column=4).value)
            if driers == "79691782&did=33556432" or driers == "79691777&did=33555432":
                single.append("5")
            else:
                single.append("0")
            driers = ""
            refresh()

    # Deleting empty tasks
    cleartasks = tasks.copy()
    for t in range(len(tasks)):
        if tasks[t][2] == "None":
            cleartasks.remove(tasks[t])
    schedule.clear()

    # keyboard()
    # if keyboard() == "ru":
    #     hotkey("alt", "shift")

    for i in range(len(cleartasks)):
        exec(f"""schedule.every().{cleartasks[i][1]}.at('{cleartasks[i][2]}').do(turn, '{(cleartasks[i][0])}','&vid=17&value={cleartasks[i][3]}')""")
        # hotkey("ctrl", "w")
    schedule.every(10).minutes.do(runschedule)
    cleartasks.clear()
    tasks.clear()

runschedule()


while True:
    schedule.run_pending()
    time.sleep(1)






