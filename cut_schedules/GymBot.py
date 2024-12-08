############################   HEADERS   ############################################

sauter_cookie = {"JSESSIONID": "b3euu8eyg51qpidr6w6qlg5b",
                 "pw": "73F0CF35AF3850221A1D641803815D686EC21A9316F6CFBFF079AE3B9BFE8DEA",
                 "BAYEUX_BROWSER": "d4f7-wxmcjfmvlu0slpxijixj162f",
                 "uid": "0"}
header ={
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-GB,en;q=0.9,en-US;q=0.8",
        "Connection": "keep-alive",
        #"Host": "192.168.250.50",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.42",
        "X-Requested-With": "XMLHttpRequest"
        }


header_alarm_A={
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "en-GB,en;q=0.9,en-US;q=0.8",
    "Connection": "keep-alive",
    "Host": "192.168.250.50",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.57",
                }



#################################   PLANTS    ###################################

plant = {
        '79691777&did=33555432': 'ПВ-1.1 ',
        '79691782&did=33556432': 'ПВ-2.1 ',
        '8388619&did=33560432':  'ПВ-1.2 ',
        '8388864&did=33557432':  'ПВ-1.5 ',
        '8388871&did=33557432':  'ПВ-1.6 ',
        '8388858&did=33557432':  'ПВ-1.7 ',
        '8388708&did=33557432':  'П-1.8  ',
        '8388676&did=33559432':  'П-1.9  ',
        '8388704&did=33557432':  'П-1.10 ',
        '8388743&did=33559432':  'П-1.11 ',
        '8388762&did=33560432':  'ПВ-2.2 ',
        '8388822&did=33561432':  'ПВ-2.3 ',
        '8388808&did=33561432':  'ПВ-2.4 ',
        '8388778&did=33560432':  'ПВ-2.5 ',
        '8388770&did=33560432':  'ПВ-2.6 ',
        '8388827&did=33561432':  'ПВ-2.7 ',
        '8388835&did=33561432':  'ПВ-2.8 ',
        '8388815&did=33561432':  'ПВ-2.9 ',
        '8388788&did=33560432':  'ПВ-2.10',
        '8388739&did=33559432':  'ПВ-2.11',
        '8388750&did=33559432':  'ПВ-2.12',
        '8388757&did=33559432':  'ПВ-2.13',
        '8388795&did=33560432':  'ПВ-2.14',
        '8388845&did=33559432':  'ПВ-2.15',
        '8388650&did=33561432':  'П-2.18 ',
        '8388672&did=33559432':  'П-2.19 ',
        '8388654&did=33561432':  'П-2.20 ',
        '8388614&did=33560432':  'П-2.21 ',
        '8388902&did=33560432':  'П-2.23 ',
        '8388851&did=33559432':  'П-2.24 ',
        '79992777&did=33554453': 'В-1.9  ',
        '79993777&did=33554453': 'B-1.38 ',
        '8388883&did=33560432':  'Бассейн'
        }

alarms_A={
        #'20971538&did=33555432':  'ПВ-1.1 ',
        # '79691782&did=33556432':  'ПВ-2.1 ',
        '12582981&did=33560432':  'ПВ-1.2 ',
        '12584064&did=33557432':  'ПВ-1.5 ',
        '12584092&did=33557432':  'ПВ-1.6 ',
        '12584036&did=33557432':  'ПВ-1.7 ',
        '12583436&did=33557432':  'П-1.8  ',
        '12583282&did=33559432':  'П-1.9  ',
        '12583415&did=33557432':  'П-1.10 ',
        '12583596&did=33559432':  'П-1.11 ',
        '12583680&did=33560432':  'ПВ-2.2 ',
        '12583904&did=33561432':  'ПВ-2.3 ',
        '12583848&did=33561432':  'ПВ-2.4 ',
        '12583736&did=33560432':  'ПВ-2.5 ',
        '12583708&did=33560432':  'ПВ-2.6 ',
        '12583960&did=33561432':  'ПВ-2.7 ',
        '12583932&did=33561432':  'ПВ-2.8 ',
        '12583876&did=33561432':  'ПВ-2.9 ',
        '12583764&did=33560432':  'ПВ-2.10',
        '12583575&did=33559432':  'ПВ-2.11',
        '12583624&did=33559432':  'ПВ-2.12',
        '12583652&did=33559432':  'ПВ-2.13',
        '12583792&did=33560432':  'ПВ-2.14',
        '12583988&did=33559432':  'ПВ-2.15',
        '12583142&did=33561432':  'П-2.18 ',
        '12583261&did=33559432':  'П-2.19 ',
        #'12583163&did=33561432':  'П-2.20 ',
        '12582953&did=33560432':  'П-2.21 ',
        # '8388902&did=33560432':   'П-2.23 ',
        '12584008&did=33559432':  'П-2.24 ',
        # # '21272522&did=33554453&vid=80': 'В-1.9.1  ',
        # '7d9992777&did=33554453':       'В-1.9.2  ',
        # '21273522&did=33554453&vid=80': 'B-1.38.1 ',
        # '21273527&did=33554453&vid=80': 'B-1.38.2 ',
        # '8388883&did=33560432':         'Бассейн'
        }

alarms_BC={
        #'20971538&did=33555432': 'ПВ-1.1 ',
        # '79691782&did=33556432': 'ПВ-2.1 ',
        '12582982&did=33560432':  'ПВ-1.2 ',
        '12584065&did=33557432':  'ПВ-1.5 ',
        "12584093&did=33557432":  'ПВ-1.6 ',
        '12584037&did=33557432':  'ПВ-1.7 ',
        '12583437&did=33557432':  'П-1.8  ',
        '12583283&did=33559432':  'П-1.9  ',
        '12583416&did=33557432':  'П-1.10 ',
        '12583597&did=33559432':  'П-1.11 ',
        '12583681&did=33560432':  'ПВ-2.2 ',
        '12583905&did=33561432':  'ПВ-2.3 ',
        '12583849&did=33561432':  'ПВ-2.4 ',
        '12583737&did=33560432':  'ПВ-2.5 ',
        '12583709&did=33560432':  'ПВ-2.6 ',
        '12583961&did=33561432':  'ПВ-2.7 ',
        '12583933&did=33561432':  'ПВ-2.8 ',
        '12583877&did=33561432':  'ПВ-2.9 ',
        '12583765&did=33560432':  'ПВ-2.10',
        '12583576&did=33559432':  'ПВ-2.11',
        '12583625&did=33559432':  'ПВ-2.12',
        '12583653&did=33559432':  'ПВ-2.13',
        '12583793&did=33560432':  'ПВ-2.14',
        '12583989&did=33559432':  'ПВ-2.15',
        '12583143&did=33561432':  'П-2.18 ',
        '12583262&did=33559432':  'П-2.19 ',
        #'12583164&did=33561432':  'П-2.20 ',
        '12582954&did=33560432':  'П-2.21 ',
        # '8388902&did=33560432':  'П-2.23 ',
        '12584009&did=33559432':  'П-2.24 ',
        # '79992777&did=33554453': 'В-1.9  ',
        # '79993777&did=33554453': 'B-1.38 ',
        # '8388883&did=33560432':  'Бассейн'
        }


driers = ["79691782&did=33556432", "79691777&did=33555432"]
swpool = ["8388883&did=33560432",]




##############################   HELPMSG   ######################################

helpmsg_halls = """
Бот управляет вентустановками:
Игрового зала.
Залов хореографии.
Раздевалок игрового зала.

Что может бот:
Запустить вентустановку. 
Изменить скорость уже работающей установки.
Остановить вентустановку. 
Узнать состояние вентустановки (работает/ скорость работы/ остановлена/ в аварии). 
Узнать заданное расписание для автоматического пуска/останова. 

Для чего нужен бот (пример):
Спортсмены не пришли в зал, вентиляция осталась включенной по расписанию. 
Позвонили на пост охраны, просят запустить вентиляцию, изменить скорость работы.
Расписание не сработало, установки остались работать.
И т.п. 

Как пользоваться ботом:
Выбрать помещение.
Выбрать действие.
Все делать при помощи кнопок.

Общие пояснения:
Бездумно кнопки нажимать не нужно (например: запустить, тут же остановить, тут же поменять скорость, выбрать другое действие, не дождавшись ответа и т.п.).
Бот будет выполнять каждую Вашу команду.
Если установка уже работает и нажать запуск на той же скорости - ничего не произойдет, установка продолжит работу.
Если в процессе работы выбрать другую скорость - установка перейдет на выбранную скорость.

Расписание работы (включение/ выключение по часам) задается с диспетчерского компьютера в п.2090.

Бот может подвисать (особенно кнопка состояние), но через пару минут все поднимется. 

С телефона тоже все работает.
"""

helpmsg_gyms = """
Бот управляет вентустановками:
Всех трех тренировочных залов.
Раздевалок тренироврчных залов.

Что может бот:
Запустить вентустановку. 
Изменить скорость уже работающей установки.
Остановить вентустановку. 
Узнать состояние вентустановки (работает/ скорость работы/ остановлена/ в аварии). 
Узнать заданное расписание для автоматического пуска/останова. 

Для чего нужен бот (пример):
Окна открыты, вентиляция работает впустую ,можно выключить. 
Позвонили, просят узнать включена вентиляция или нет, запустить, изменить скорость работы.
Расписание не сработало, установки остались работать.
И т.п. 

Как пользоваться ботом:
Выбрать помещение.
Выбрать действие.
Все делать при помощи кнопок.

Общие пояснения:
Бездумно кнопки нажимать не нужно (например: запустить, тут же остановить, тут же поменять скорость, выбрать другое действие, не дождавшись ответа и т.п.).
Бот будет выполнять каждую Вашу команду.
Если установка уже работает и нажать запуск на той же скорости - ничего не произойдет, установка продолжит работу.
Если в процессе работы выбрать другую скорость - установка перейдет на выбранную скорость.
Кнопка "Тренажерный зал" подразумевает только большой первый Г-образный тренажерный зал,
остальные два зала - это кнопка "Зал Ёги"

Расписание работы (включение/ выключение по часам) задается с диспетчерского компьютера в п.2090.

Бот может подвисать (особенно кнопка состояние), но через пару минут все поднимется. 

С телефона и ноутбука все одинаково работает.
"""




##########################   TELEGRAMTOKEN   ####################################

telegramtoken = "6182428747:AAF9uTc4GsJ_TpB3p-dNwq0hwrMEwF4Bm5Y"
telegramtoken_venthalls = "6326141804:AAF8XfDaHIShYwmwDU-Q9X2D3c95zZhU0Z0"
telegramtoken_ventgyms  = "6358702276:AAGBzDWfxR_7wh3Za4NEe_MCVCjQaD1PSYg"


telegram_users = {"Mikhail": "1140621075",
                  "Michael":  "5740110040",
                  "Roman":    "1565146153",
                  "Alexandr": "981504594",
                  "Nikita":   "397472880",
                  "Andrey":   "2116310630",
                  "Yury":     "1969627530",
                  "Euheny":   "1016661717",
                }

botHalls_users = {"Mikhail": "1140621075",
                  "Michael":  "5740110040",
                  "Roman":    "1565146153",
                  'Security': "6608329301",
                  'Andrey' :  "2116310630",
                  }
botGyms_users = { "Mikhail":   "1140621075",
                  "Michael":  "5740110040",
                  "Roman":    "1565146153",
                  "Aleksandr": "1094527985",
                 }


#########################   TELEGRAMBOTALARMS   #################################

import telebot
# from telegram.telegramtokens import telegramtoken, telegram_users
from duties import onduties

bot = telebot.TeleBot(telegramtoken)


def to_telegram(msg):
    """
    telegram()     - отправка события в мессенджер Telegram
    msg            - текст сообщения
    bot            - объект класса TeleBot для работы с Telegram
    telegramtoken  - токен мессенджера Telegram
    telegramusers  - идентификаторы получателей сообщений в Telegram
    """
    bot = telebot.TeleBot(telegramtoken)
    for i in telegram_users.values():
        bot.send_message(i, msg)

    # получаем ключ (имя получателя) для словаря с именами и id
    # worker = onduties()

    # отправляем сообщение тому, кто сейчас на смене
    # bot.send_message(chat_id_workers[worker], msg)


# if __name__ =="__main__":
#     bot = telebot.TeleBot(telegramtoken)
#     msg="Авария на объекте, просьба относиться с пониманием!"
#     bot.send_message(telegram_users["Mikhail"], msg)
#     # for i in telegram_users.values():
#     #     bot.send_message(i, msg)
######################################################################################################################################333

################################    SHCEDULE.PY          ########################################
import schedule
import time
from logs import log, readlogs
from excels import readschedule
from request import switch, getalarms
from plants import alarms_A, alarms_BC


print("РАБОТАЕТ УПРАВЛЕНИЕ ОБОРУДОВАНИЕМ ПО РАСПИСАНИЮ, НЕ ЗАКРЫВАЙТЕ ЭТО ОКНО")

tasks = []


def turn(get_plant, par):
    """
    turn()       - выполнение действия (включить, выключить и т.п.)
    get_plant    - параметр для формирования url, определяющий код установки
    par          - параметр для формирования url, определяющий действие (0 -стоп, 1 пуск и т.п.)
    log()        - запись отправленных команд в лог файлы
    switch()     - выполнение запроса на сервер при помощи библиотеки requests
    """
    log(get_plant, par)
    switch(get_plant, par)


def runschedule():
    """
    runschedule()    - получение расписания, удаление пустых значений, компоновка задач и запуск действий по расписанию.
    readschedule()   - чтение расписания из excel файла принимает пустой список tasks, возвращает заполненный список
    cleartasks       - список, где не будет пустых расписаний
    tasks            - полный список расписаний, в т.ч. пустых
    clear('cleared') - очистка предыдущего schedule c тегом 'cleared',
                       т.к. периодически происходит его чтение и формирование заново
    exec()           - автоматическая компоновка задачи для функции schedule
    """
    clearlogs = readschedule(tasks)
    cleartasks = clearlogs.copy()
    for t in range(len(tasks)):
        if tasks[t][2] == "None":
            cleartasks.remove(tasks[t])
    schedule.clear('cleared')
    for i in range(len(cleartasks)):
        exec(f"schedule.every().{cleartasks[i][1]}"
             f".at('{cleartasks[i][2]}')"
             f".do(turn,'{cleartasks[i][0]}','&vid=17&value={cleartasks[i][3]}')"
             f".tag('cleared')")
    readlogs(cleartasks)
    cleartasks.clear()
    tasks.clear()


schedule.every(2).minutes.do(runschedule)
schedule.every(13).minutes.do(getalarms, alarms_dict=alarms_A,  column_number=4, alarm_text='Авария класса А')
schedule.every(23).minutes.do(getalarms, alarms_dict=alarms_BC, column_number=5, alarm_text='Авария класса B,C')


while True:
    schedule.run_pending()
    time.sleep(1)
#################################################################################################################



###################################### LOGS.PY   ################################################################
import datetime
# from plants import plant, driers, swpool
from pywinauto.application import Application


class Textjob:
    pathcur = "./logging/log_scheduling.txt"
    pathall = "./logging/alllogs.txt"
    readlog = "./logging/readlogs.txt"
    def __init__(self, path, flag):
        self.path = path
        self.flag = flag
    def makelog(self, data, n=""):
        f = open(self.path, self.flag)
        f.write("".join(data) + n)
        f.close()

def log(plantcode, acting):
    """
    log()       -  основная функция логгирования"
    plantcode   -  код установки
    acting      -  действие в читаемом виде
    close()     -  закрытие лог файла, если на момент записи в него он открыт
    logwrite    -  текст задания, пример:   "17-05-2023  20:30  ПВ-2.6   Cтоп"
    """
    close()
    logwrite = [datetime.datetime.now().strftime("%d-%m-%Y  %H:%M  "), plant[f"{plantcode}"], act(plantcode, acting)]
    l = Textjob(Textjob.pathall, "a")
    l.makelog(logwrite, "\n")
    sort()


def close():
    """
    close()  - закрытие лог файла, если на момент записи в него он открыт
    try      - попытка найти открытый файл с заголовком "log_scheduling" и закрыть его
    """
    ntp = Application()
    try:
        ntp.connect(title_re="log_scheduling")
        ntp.window().close()
    except:
        pass

def act (singlecode, whattodo):
    """
    act()       - определение действия
    dryers      - список с кодами осушителей
    swpool      - список с кодом бассейна
    action      - читаемое действие для записи в лог
    a           - последний символ запроса (закодированное действие)
    """
    action = ""
    a = whattodo[-1:]
    if   a == "1": action = "  Пуск на низкой скорости"
    elif a == "2": action = "  Пуск на высокой скорости"
    elif a == "0": action = "  Cтоп"
    # If swimming pool
    if singlecode in swpool:
        if   a == "0": action = "  Выключение подсветки"
        elif a == "2": action = "  Включение желтой подсветки"
        elif a == "1": action = "  Включение синей подсветки"
    elif singlecode in driers:
        if   a == "5": action = "  Стоп"
        elif a == "1": action = "  Пуск в режиме хоккей"
        elif a == "2": action = "  Пуск в режиме фигурное катание"
    return action

def writelog(parttasks, partlogs):
    """
    writelog()  - запись в лог файл команд за заданный период
    abs()       - необходима для корректной вставки пустой строки при наступлении  нового месяца
    logtasks    - список для записи в файл log_scheduling.txt
    alllogs     - список для записи в файл alllogs.txt
    parttasks   - список списков заданий за короткий заданный период
    partlogs    - список списков заданий за длинный заданный период
    в первом цикле также вставка пустой строки между разными датами
    """
    logtasks=[]
    alllogs =[]
    for c in range(len(parttasks)):
        if c >= 1 :
            if abs(int(parttasks[c][0:2])-int(parttasks[c-1][0:2])) >= 1:
               logtasks.append('\n')
        logtasks.append(f'{"".join(parttasks[c])}\n')
    for c in range(len(partlogs)):
        alllogs.append(f'{"".join(partlogs[c])}\n')
    f = Textjob(Textjob.pathcur, "w")
    d = Textjob(Textjob.pathall, "w")
    d.makelog(alllogs)
    f.makelog(logtasks)


def sort():
    """
    sort()    - сортировка команд за заданный период для записи в лог-файл
    current   - список из всех записей в alllogs, разделенный по символу переноса
    parttasks - список списков заданий за короткий заданный период
    partlogs  - список списков заданий за длинный заданный период
    в цикле проверка на предмет выдавать записи за несколько дней и ограничение всего периода записей"""
    f = open(Textjob.pathall, "r")
    current = f.read().split("\n")
    parttasks = []
    partlogs  = []
    rightnow = datetime.datetime.now()
    for i in range(len(current)-1):
        logtime = datetime.datetime.strptime(current[i][0:16], "%d-%m-%Y  %H:%M")
        if rightnow - datetime.timedelta(days=1) <= logtime:
            parttasks.append(current[i])
        if rightnow - datetime.timedelta(days=10) < logtime:
            partlogs.append(current[i])
    f.close()
    writelog(parttasks, partlogs)

days={   "sunday": "Воскресенье",
         "monday": "Понедельник",
        "tuesday": "Вторник    ",
      "wednesday": "Среда      ",
       "thursday": "Четверг    ",
         "friday": "Пятница    ",
       "saturday": "Суббота    "}

def readlogs(logs_read):
    """
    readlogs()   - запись прочитанных заданий с расписаниями в лог файл readlogs.txt
    logs_read    - список списков , где нет пустых расписаний
    reads        - список прочитанных заданий за 7 дней для записи в файл readlogs.txt
    ttdys        - список прочитанных заданий за 2 дня для записи в файл readlogs.txt
    today        - текущий день недели в формате одной цифры
    yrday        - вчерашний день недели в формате одной цифры
    str_today    - текущий день недели
    str_yrday    - вчерашний день недели
    readlog      - путь к файлу readlogs.txt
    """
    reads = []
    ttdys = []
    today = datetime.datetime.now().date()
    yrday = today - datetime.timedelta(days=1)
    tmday = today + datetime.timedelta(days=1)
    str_today = today.strftime("%A").lower()
    str_yrday = yrday.strftime("%A").lower()
    str_tmday = tmday.strftime("%A").lower()
    for i in logs_read:
        if i[1] == str_today  or i[1] == str_tmday: #or i[1] == str_yrday:
            ttdys.append(i)
    for i in ttdys:
        reads.append(f"{plant[i[0]]}   {''.join(days[i[1]])}   {''.join(i[2][0:5])} {(act(i[0], i[3]))} \n")
    t = Textjob(Textjob.readlog, "w")
    t.makelog(reads)



# if __name__ == "__main__":
#     pass
    # pass
    # sort()
    # log('8388762&did=33560432',"0")
    # act('8388762&did=33560432',"0")
################################################################################################################

############################################################################################################


#############################################################################################################################################


#####################################################REQUEST.PY###############################################################################
from openpyxl import load_workbook
from datetime import datetime
from logs import sort, Textjob
from plants import driers
from telegram.TelegramBotAlarms import to_telegram
from viber.ViberSet import to_viber

schedule_book = "./excel/Расписание.xlsm"
status_book   = "./excel/Cостояние.xlsx"
single = []


def refresh(tasks):
    """
    refresh()        - очистка списка single и пополнение списка tasks
    tasks            - полный список расписаний, в т.ч. пустых
    single           - по итогу список с одной задачей
                       (установка, день недели, что сделать, во сколько сделать)
    replace()        - исправление не корректного чтения
    """
    tasks.append(single.copy())
    single.clear()

def refreeze(row, lis, book):
    """
    refreeze()       - чтение кода установки из файла excel
    row              - номер ряда в файле excel
    lis              - список single
    book             - рабочая книга Расписание.xlsm
    plancod          - прочтенное из файла excel значение кода установки
    """
    plancod = (str(book.active.cell(row=row, column=4).value))
    lis.append(plancod)
    return plancod

def drier(row, lis, book):
    """
    drier()          - проверка не является ли установка осушителем
                       (имеет нетиповые параметры управления)
    driers           - коды установок, где стопу соответствует параметр "5" , а не "0"
    row              - номер ряда в файле excel
    lis              - список single
    book             - рабочая книга Расписание.xlsm
    """
    dry = str(book.active.cell(row=row, column=4).value)
    if dry in driers:
        lis.append("5")
    else:
        lis.append("0")

def readschedule(tasks):
    """
    readschedule()   - чтение расписания из excel файла принимает пустой список tasks, возвращает заполненный список
    tasks            - полный список расписаний, в т.ч. пустых
    workbook         - работа с рабочей книгой excel
    refresh()        - предобработка заданий с расписанием
    """
    workbook = load_workbook(schedule_book)
    for j in range(53, 383, 10):
        for k in range(1, 8):
            refreeze(j, single, workbook)
            for s in [2, 5, 3]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 6]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            drier(j, single, workbook)
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 8, 10]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            refresh(tasks)
            refreeze(j, single, workbook)
            for s in [2, 9]:
                single.append(str(workbook.active.cell(row=j + k, column=s).value))
            drier(j, single, workbook)
            refresh(tasks)
    return tasks

def writestatus(i, plant, alarm, column):
    """
    writestatus()  - запись сведений об аварии в файл Состояние.xlsx
    i              - сдвиг номера строки для внесения записей в файл Состояние.xlsx
    plant          - имя установки, заносимое в файл Состояние.xlsx (например "ПВ-1.6")
    alarm          - запись об аварии, заносимая в файл Состояние.xlsx (например "Авария класса А")
    column         - номер колонки для записи в файл Состояние.xlsx
    statusbook     - объект для работы с файлом Состояние.xlsx
    date_now       - текущая дата, заносимая в файл Состояние.xlsx
    time_now       - текущее время, заносимая в файл Состояние.xlsx
    alarm_happen   - сторка типа "04-06-2023  11:00  ПВ-2.8   Авария класса А"
    makelog()      - занесение события об аварии в лог файл alllogs.txt
    sort()         - занесение события об аварии в лог файл log_scheduling.txt
    to_telegram()  - отправка сообщения об аварии в чат бот в Телеграм
    to_viber()     - отправка сообщения об аварии в чат бот в Viber
    """
    statusbook = load_workbook(status_book)
    date_now = datetime.now().strftime("%d-%m-%Y")
    time_now = datetime.now().strftime("%H:%M")
    statusbook.active.cell(row=3+i, column=1).value = date_now
    statusbook.active.cell(row=3+i, column=2).value = time_now
    statusbook.active.cell(row=3+i, column=3).value = plant
    if statusbook.active.cell(row=3+i, column=column).value != alarm:
        alarm_happen = [f"{date_now}  ", f"{time_now}  ", f"{plant}  ",  alarm]
        m = Textjob(Textjob.pathall, "a")
        m.makelog(alarm_happen, "\n")
        sort()
        logmsg = "\n".join(alarm_happen[2:])
        to_telegram(logmsg)
        to_viber(logmsg)
    statusbook.active.cell(row=3 + i, column=column).value = alarm
    statusbook.save(status_book)
##############################################################################################################################################




##################################TELEGRABBOTGYMS.PY########################################################################################
# import telebot
# from telegramtokens import telegramtoken_ventgyms, botGyms_users
from telebot import types
# import requests
# import time
from datetime import datetime
# from headers import header, header_alarm_A, sauter_cookie
# from plants import alarms_A, alarms_BC
# from helpmsg import helpmsg_gyms

# Инициализация бота
gym_bot = telebot.TeleBot(telegramtoken_ventgyms)

# Наименования оборудования
gym_places = {"Игровая комната": "ПВ-2.9",
          "Раздевалки залов": "ПВ-2.15",
          "Тренажерный зал": "ПВ-2.12",
          "Зал Ёги": "ПВ-2.11"}

# Парметры для запроса к оборудованию
gym_all_plants = {"ПВ-2.11": "8388739&did=33559432",
              "ПВ-2.12": "8388750&did=33559432",
              "ПВ-2.15": "8388845&did=33559432",
              "ПВ-2.9": "8388815&did=33561432", }

# Коды состояний работы
gym_states = {"2": "работает на высокой скорости",
          "1": "работает на низкой скорости",
          "0": "остановлена",
          "3": "работает на низкой скорости", }

# Тексты команд на кнопках
gym_starts = ["Запуск  " + x for x in gym_places.values()]
gym_stops = ["Останов  " + x for x in gym_places.values()]
gym_curstates = ["Состояние  " + x for x in gym_places.values()]
gym_scheds = ["Расписание  " + x for x in gym_places.values()]


def gym_if_root(permit):
    """
    Ограничение доступа к боту

    :param permit: функция, для которой будет применяться ограничение прав доступа
    :return: функция для сравнения входящего id со списком пользователей
    """
    def check_root(message):
        """
        Ограничение прав доступа

        :param message: входной параметр для определения id пользователя
        """
        uid = str(message.chat.id)
        if uid in botGyms_users.values():
            permit(message)
        else:
            sms(uid, "У Вас нет прав доступа к этому боту")
    return check_root


def gym_alrm_params(alrm_dict):
    """ Преобразование с целью получения параметров для запроса

    :param alrm_dict: словарь с параметрами (...'12583575&did=33559432':  'ПВ-2.9 ',)
    :return:          словарь с корректными параметрами (...'ПВ-2.9': '12583575&did=33559432')
    """
    return {v.replace(" ", ""): k for k, v in alrm_dict.items()
            if v.replace(" ", "") in all_plants.keys()}


@bot.callback_query_handler(func=lambda callback: callback.data in ['1', '2'])
def gym_check_speed(callback):
    """
    Обработчик сообщения с инлайн кнопками

    :param callback: объект, содержащий в т.ч. инфо о нажатой кнопке
    """
    uid = callback.message.chat.id
    tex = callback.message.text.replace("Выберите скорость работы", "Запуск ")
    sms(uid,  "Стартуем.... ", 4)
    switch_plant(callback.message, tex, callback.data, "Запуск")


@bot.message_handler(commands=['help'])
@gym_if_root
def gym_hepl(message):
    """
    Обработчик сообщения "help"

    :param message: объект "сообщение" (содержит в т.ч. текст сообщения)
    """
    uid = message.chat.id
    sms(uid, helpmsg_gyms)
    user_action(message, tex="Инструкция")


@bot.message_handler(commands=['start'])
@gym_if_root
def start(message):
    """
    Обработчик сообщений, поступающих после команды "start"

    :param message: объект "сообщение" (содержит в т.ч. текст сообщения)
    """
    uid = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("Игровая комната")
    btn2 = types.KeyboardButton("Раздевалки залов")
    btn3 = types.KeyboardButton("Тренажерный зал")
    btn4 = types.KeyboardButton("Зал Ёги")
    markup.add(btn3, btn4)
    markup.add(btn2)
    markup.add(btn1)
    bot.send_message(uid, "Выберите помещение", reply_markup=markup)


@bot.message_handler(content_types=['text'])
@if_root
def func(message):
    """
    Обработчик текстовых сообщений (основная логика)

    :param message: объект "сообщение"
    """
    uid = message.chat.id
    msg = message.text
    pv = msg.index("ПВ") if "ПВ" in msg else 0
    user_action(message, msg)
    # Главное меню
    if msg == "Главное меню":
        start(message)
    # Информация
    elif msg in places.keys():
        sms(uid, f"{msg}\nобслуживается вентустановкой  {places[msg]}", 1)
        reply(message, msg)
    # Расписание
    elif msg in scheds:
        sms(uid, f"Ждите, сейчас узнаем ...", 4)
        fil = open("../logging/readlogs.txt", "r")
        sts = []
        for i in fil.read().split("\n"):
            if msg[pv:] in i:
                sts.append(i.replace(f"{msg[pv:]}    ", ""))
        prn = "\n".join(sts).replace("\n", "\n\n").replace("0   ", "0\n")
        if prn == "":
            prn = "не задано"
        sms(uid, f'{msg}\nна эти дни:\n\n{prn}', 1)
        user_action(message, tex=f'{msg} на эти дни: {prn[:9]}...')
        sms(uid)
    # Состояние
    elif msg in curstates:
        sms(uid, f"Ждите, идет опрос ...",)
        state = get_state(msg[pv:])
        # alarmA = get_alarm(msg[pv:], alrm_params(alarms_BC), 'Авария класса ВС')
        # alarmB = get_alarm(msg[pv:], alrm_params(alarms_A),  'Авария класса А')
        sms(uid, f"В текущий момент установка {msg[pv:]}  {state}.\n")  # {alarmA} {alarmB}")
        user_action(message, tex=f"Установка {msg[pv:]} {state}")
        sms(uid)
    # Запуск
    elif msg in starts:
        markup = types.InlineKeyboardMarkup()
        button2 = types.InlineKeyboardButton("Низкая",  callback_data="1")
        button1 = types.InlineKeyboardButton("Высокая", callback_data="2")
        markup.add(button2, button1)
        spd = "Выберите скорость работы"
        bot.send_message(message.chat.id, f"{spd} {msg[pv:]}", reply_markup=markup)
    # Останов
    elif msg in stops:
        sms(uid, "Останавливаемся....", 3)
        switch_plant(message, msg, "0", "Останов")
    # Иное
    else:
        start(message)


def sms(uid, t="Выберите действие", s=0):
    """
    Отправка текстовых сообщений в бот

    :param uid: идентификатор адресата сообщения (chat.id)
    :param t:   текст сообщения (по умолчанию - "Выберите действие")
    :param s:   задержка времени после отправки сообщения
    """
    bot.send_message(uid, t)
    time.sleep(s)


def switch_plant(message, msg, act, action):
    """
    Подготовка к выполнению действия над вентустановкой

    :param message: объект "сообщение", содержит в т.ч. текст полученного сообщения
    :param msg:     текст полученного сообщения
    :param act:     параметр, определяющий действие (0-стоп, 1-запуск на низкой и т.п.)
    :param action:  текст, соответствующий действию ("Запуск", "Останов" и т.п)
    :cod:           идентификатор установки (например "8388750&did=33559432")
    :pv:            местораположение наименования установки в сообщении
    """
    cod = all_plants[msg.replace(f"{action}  ", "")]
    pv = msg.index("ПВ")
    sw = do_switch(cod, act, msg[pv:])
    bot.send_message(message.chat.id, f"{msg} {sw}")
    user_action(message, f"{msg} {sw}")


def do_switch(g, p, plt):
    """
    Выполнение действия над вентустановкой

    :param g:    параметр, определяющий адрес установки для формирования строки запроса
    :param p:    параметр, определяющий действие (0-стоп, 1-запуск на низкой и т.п.)
    :param plt:  наименование установки (ПВ-2.9, ПВ-2.11 и т.п.)
    """
    stmsg = "не выполнен. "
    url = f"http://192.168.250.50/ajaxjson/bac/setValue?pid=85&oid={g}&vid=17&value={p}"
# if check_alarm(plt):
#     stmsg = stmsg + "Авария класса А"
# else:
    r = requests.get(url, headers=header, cookies=sauter_cookie)
    time.sleep(3)
    if '"message":"Value was successfully written"' in r.text:
        stmsg = "выполнен успешно."
    return stmsg


def get_state(plt):
    """
    Получение состояния вентустановки

    :param plt:  наименование установки (ПВ-2.9, ПВ-2.11 и т.п.)
    :return:     полученное состояние вентустановки (работает, остановено, в аварии и т.п.)
    """
    url = f"http://192.168.250.50/svo/details/?oid={all_plants[plt]}&vid=17&mode=cached"
    resp = requests.get(url, headers=header, cookies=sauter_cookie)
    time.sleep(2)
    rsp = resp.text
    num = rsp[rsp.index('<tr data-pid="85">')+18:]
    end = num[:num.index('</tr>')]
    stt = end[end.index("property-value")+16]
    return states[stt]


def get_alarm(plt, dic, txt):
    """
    Получения сведений о возможной аварии вентустановки

    :param plt:   наименование вентустановки (например "ПВ-2.9")
    :param dic:   словарь с параметрами для строки запроса
    :param txt:   текст сообщения об аварии
    :return:      текст сообщения об аварии (например "Авария класса А")
    """
    url = f"http://192.168.250.50/svo/details/update?oid={dic[plt]}&vid=17&mode=cached"
    resp = requests.get(url, headers=header, cookies=sauter_cookie)
    time.sleep(2)
    almsg = ""
    if "Alarm: true" in resp.text:
        if 'title="Fault: true"' not in resp.text:
            almsg = txt
    return almsg


def check_alarm(plt):
    """
    Проверка на наличие аварии перед пуском

    :param plt: наименование вентустановки (например "ПВ-2.9")
    :return:    True - есть авария, False - нет аварии
    """
    prs = alrm_params(alarms_A)
    alm = 'Авария класса А'
    return True if get_alarm(plt, prs, alm) == alm else False


# @if_root
def reply(message, place=""):
    """
    Обработчик сообщения "выберите действие"

    :param message: объект "сообщение" (например "Состояние ПВ-2.9")
    :param place:   наименование помещения (например "Игровая комната")
    """
    answer = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton(f"Состояние  {places[place]}")
    button2 = types.KeyboardButton(f"Расписание  {places[place]}")
    button3 = types.KeyboardButton(f"Запуск  {places[place]}")
    button4 = types.KeyboardButton(f"Останов  {places[place]}")
    button5 = types.KeyboardButton("Главное меню")
    answer.add(button1, button2)
    answer.add(button3, button4)
    answer.add(button5)
    bot.send_message(message.chat.id, "Выберите действие".format(message.from_user), reply_markup=answer)


def user_action(message, tex=""):
    """
    Вывод действий пользователя и ответов в консоль

    :param message:  объект "сообщение"
    :param tex:      отображаемый текст
    """
    uid = message.chat.id
    fst = message.from_user.first_name
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(now, uid, fst, tex, sep="  ")



bot.infinity_polling(none_stop=True,
                      timeout=180,
                      long_polling_timeout=180,
                      allowed_updates=['message', 'callback_query'])
#
# while True:
#     try:
#         bot.polling()
#     except:
#         print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   ..........  сбой соединения")
#         time.sleep(10)
####################################################################################################################################################3


####################################TELEGRAMBOTHALLS.PY###########################################################################3
import telebot
# from telegramtokens import telegramtoken_venthalls, botHalls_users
from telebot import types
import requests
import time
# from headers import header, header_alarm_A, sauter_cookie
# from plants import alarms_A, alarms_BC
# from helpmsg import helpmsg_halls

bot = telebot.TeleBot(telegramtoken_venthalls)

places = {"Игровой зал": "ПВ-2.7, ПВ-2.8",
          "Раздевалки игрового зала": "ПВ-2.4",
          "Зал хореографии 2015": "ПВ-2.5",
          "Зал хореографии 2041": "ПВ-2.6"}

all_plants = {"ПВ-2.4": "8388808&did=33561432",
              "ПВ-2.5": "8388778&did=33560432",
              "ПВ-2.6": "8388770&did=33560432",
              "ПВ-2.7": "8388835&did=33561432",
              "ПВ-2.8": "8388827&did=33561432",
              }

states = {"2": "работает на высокой скорости",
          "1": "работает на низкой скорости",
          "0":  "остановлена"}

starts = ["Запуск  " + x for x in places.values()]
stops = ["Останов  " + x for x in places.values()]
curstates = ["Состояние  " + x for x in places.values()]
scheds = ["Расписание  " + x for x in places.values()]

rev_alarms_BC = {v[:6]: k for k, v in alarms_BC.items() if v[:6] in all_plants.keys()}
rev_alarms_A = {v[:6]: k for k, v in alarms_A.items() if v[:6] in all_plants.keys()}


@bot.message_handler(commands=['help'])
def start(message):
    m = message.chat.id
    if root(m):
        sms(m, helpmsg_halls)
    else:
        no_root(m)


@bot.message_handler(commands=['start'])
def start(message):
    m = message.chat.id
    if root(m):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("Игровой зал")
        btn2 = types.KeyboardButton("Раздевалки игрового зала")
        btn3 = types.KeyboardButton("Зал хореографии 2015")
        btn4 = types.KeyboardButton("Зал хореографии 2041")
        markup.add(btn3, btn4)
        markup.add(btn2)
        markup.add(btn1)
        bot.send_message(m, "Выберите помещение", reply_markup=markup)
    else:
        no_root(m)


@bot.callback_query_handler(func=lambda callback: callback.data in ['1', '2'])
def check_speed(callback):
    m = callback.message.chat.id
    if root(m):
        tex = callback.message.text.replace("Выберите скорость работы", "Запуск ")
        sms(m,  "Стартуем.... ", 4)
        switch_plant(callback.message, tex, callback.data, "Запуск")
    else:
        no_root(m)


@bot.message_handler(content_types=['text'])
def func(message):
    m = message.chat.id
    if root(m):
        msg = message.text
        if msg == "Главное меню":
            start(message)

        # Информация
        elif msg in places.keys():
            vor = "обслуживает вентустановка"
            if "Игровой" in msg:
                vor = "обслуживают вентустановки"
            sms(m, f"{msg}\n{vor}  {places[msg]}", 1)
            reply(message, msg)

        # Расписание
        elif msg in scheds:
            sms(m, f"Ждите, сейчас узнаем ...", 4)
            fil = open("../logging/readlogs.txt", "r")
            sts = []
            for i in fil.read().split("\n"):
                if msg[-6:] in i:
                    sts.append(i.replace(f"{msg[-6:]}    ", ""))
            prn = "\n".join(sts).replace("\n", "\n\n").replace("0   ", "0\n")
            if prn == "":
                prn = "не задано"
            sms(m, f'{msg}\nна эти дни:\n\n{prn}', 1)
            sms(m)

        # Состояние
        elif msg in curstates:
            sms(m, f"Ждите, идет опрос ...", 2)
            sms(m, f"В текущий момент установка"
                   f" {msg[-6:]} {get_state(msg[-6:])}. "
                   f"{get_alarm(msg[-6:], rev_alarms_BC, 'Авария класса ВС')}"
                   f"{get_alarm(msg[-6:], rev_alarms_A,  'Авария класса А')}", 2)
            if "ПВ-2.7" in msg:
                sms(m, f"Установка {msg[-14:-8]} {get_state(msg[-14:-8])}."
                       f"{get_alarm(msg[-14:-8], rev_alarms_BC, 'Авария класса ВС')}"
                       f"{get_alarm(msg[-14:-8], rev_alarms_A,  'Авария класса А')}", 2)
            sms(m)

        # Запуск
        elif msg in starts:
            markup = types.InlineKeyboardMarkup()
            button2 = types.InlineKeyboardButton("Низкая", callback_data="1")
            button1 = types.InlineKeyboardButton("Высокая", callback_data="2")
            markup.add(button2, button1)
            mes = "Выберите скорость работы"
            if "ПВ-2.7" in msg:
                bot.send_message(message.chat.id, f"{mes} {msg[-14:]}", reply_markup=markup)
            else:
                bot.send_message(message.chat.id, f"{mes} {msg[-6:]}", reply_markup=markup)

        # Останов
        elif msg in stops:
            p = "0"
            sms(m, "Останавливаемся....", 3)
            switch_plant(message, msg, p, "Останов")

        # Иное
        else:
            sms(m, "Что за команда, не понял?", 3)
            sms(m, "Чувак, здесь не надо набирать текст \nПросто жмем кнопки", 3)
            sms(m, "Идем на главную")
            start(message)
    else:
        no_root(m)


def sms(m, t="Выберите действие", s=0):
    """
    Отправка текстовых сообщений в бот
    :param m:
    :param t:
    :param s:
    :return:
    """

    bot.send_message(m, t)
    time.sleep(s)


def switch_plant(message, msg, p, action):
    m = message.chat.id
    if "ПВ-2.7, ПВ-2.8" in msg:
        ssg = f"{action}  ПВ-2.7"
        sms(m, ssg)
        g = all_plants[ssg.replace(f"{action}  ", "")]
        sms(m, f"{ssg} {do_switch(g, p, 'ПВ-2.7')}")
        psg = f"{action}  ПВ-2.8"
        sms(m, psg)
        g = all_plants[psg.replace(f"{action}  ", "")]
        sms(m, f"{psg} {do_switch(g, p, 'ПВ-2.8')}")
    else:
        g = all_plants[msg.replace(f"{action}  ", "")]
        bot.send_message(message.chat.id, f"{msg} {do_switch(g, p, msg[-6:])}")


def do_switch(g, p, plt):
    stmsg = "не выполнен.\n"
    url = f"http://192.168.250.50/ajaxjson/bac/setValue?pid=85&oid={g}&vid=17&value={p}"
    if check_alarm(plt):
        stmsg = stmsg + "Авария класса А"
    else:
        r = requests.get(url, headers=header_alarm_A, cookies=sauter_cookie)
        time.sleep(4)
        if '"message":"Value was successfully written"' in r.text:
            stmsg = "выполнен успешно.\n "
    return stmsg


def get_state(pl):
    url = f"http://192.168.250.50/svo/details/?oid={all_plants[pl]}&vid=17&mode=cached"
    resp = requests.get(url, headers=header_alarm_A, cookies=sauter_cookie)
    time.sleep(3)
    r = resp.text
    n = r[r.index('<tr data-pid="85">')+18:]
    e = n[:n.index('</tr>')]
    g = e[e.index("property-value")+16]
    return states[g]


def get_alarm(pl, dic, txt):
    url = f"http://192.168.250.50/svo/details/update?oid={dic[pl]}&vid=17&mode=cached"
    resp = requests.get(url, headers=header, cookies=sauter_cookie)
    time.sleep(3)
    almsg = ""
    if "Alarm: true" in resp.text:
        if 'title="Fault: true"' not in resp.text:
            almsg = txt
    return almsg


def check_alarm(pl):
    """
    check_alarm()  - Проверка отсутствия аварии класса А перед запуском
    :param pl:     -
    :return:
    """
    alm = 'Авария класса А'
    if get_alarm(pl, rev_alarms_A, alm) == alm:
        return True
    return False


def reply(message, place=""):
    answer = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton(f"Состояние  {places[place]}")
    button2 = types.KeyboardButton(f"Расписание  {places[place]}")
    button3 = types.KeyboardButton(f"Запуск  {places[place]}")
    button4 = types.KeyboardButton(f"Останов  {places[place]}")
    button5 = types.KeyboardButton("Главное меню")
    answer.add(button1, button2)
    answer.add(button3, button4)
    answer.add(button5)
    bot.send_message(message.chat.id, "Выберите действие".format(message.from_user), reply_markup=answer)


def root(m):
    return True if str(m) in botHalls_users.values() else False


def no_root(m):
    bot.send_message(m, "У Вас нет прав доступа к этому боту")


try:
    bot.infinity_polling(none_stop=True, timeout=180, long_polling_timeout=180, allowed_updates=['message', 'callback_query'])
except Exception as er:
    time.sleep(3)
    print(er)
    pass
##################################################################################################################################

#########################################VIBERBOT.PY###################################################################
vibertoken = "5140698d1667e793-1e3d51e47aac03f2-dafdbd02d7d47bec"
viberavatar = "C:/Users/BMS/projects/schedules/logo.jpg"
viberbotname = "Вентиляция ФСК"
viberwebhook = "https://d60d-37-45-171-18.ngrok-free.app"


viber_users = {"Mikhail":  "EqHRh78MtyO9Ndjbr7BAkg==",
               "Michael":  "FBNFWql3KAXMAlnre7BdPg==",
               "Vladimir": "T4Zhj4CMqhfVeU+7Kj3c4g==",
               "Dmitry":   "ee+ZjUdnWEYCwO5MvbwqKw==",
               "Artur":    "P9KspqPKHWrndDaDHHigpQ=="}
##########################################################################################################################


#########################################VIBERBOT.PY#######################################################################
from viberbot import Api
from viberbot.api.bot_configuration import BotConfiguration
from viberbot.api.viber_requests import (
    ViberMessageRequest,
    ViberConversationStartedRequest)
from viber.vibertokens import viberbotname, vibertoken, viberavatar
from flask import Flask, request, Response


app = Flask(__name__)

bot_config = BotConfiguration(name=viberbotname, avatar=viberavatar, auth_token=vibertoken)
viber = Api(bot_config)

@app.route("/", methods=["POST"])
def incoming():
    viber_request = viber.parse_request(request.get_data())
    if isinstance(viber_request, ViberMessageRequest):
        message = viber_request.message
        viber.send_messages(viber_request.sender.id, [message])
    elif isinstance(viber_request, ViberConversationStartedRequest):
        print(viber_request.user.id)
    return Response(status=200)


#if __name__ == "__main__":
app.run(host="0.0.0.0", port =8443, debug=True )
##############################################################################################################################3